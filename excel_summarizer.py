#!/usr/bin/env python3

#
# excel_summarizer.py
#
# Date    : 2024-04-29
# Author  : Hirotoshi FUJIBE
# History :
#
# Copyright (c) 2024 Hirotoshi FUJIBE
#

"""
Usage:

    Python.exe excel_summarizer.py

Options:

    -h
    --help
        Print this message and exit.
"""

# Import Libraries
import os
import sys
import getopt
import shutil
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side

# Input, Output
IN_DIR = '.\\input'
OUT_DIR = '.\\output'
# IN_SRC_ROOT = 'D:\\Developments\\PyCharmProjects\\tool-excel_summarize\\input'  # noqa
IN_SRC_ROOT = '.\\input'
IN_SRC_RELATIVE = '\\excel'
IN_EXCEL = IN_DIR + '\\excel_summary_template.xlsx'
OUT_EXCEL = OUT_DIR + '\\excel_summary.xlsx'
IN_SHEET = 'Data Table'
OUT_SHEET = 'Summary of Data Tables'
EXTEND = '.xlsx'
OUT_DEBUG = OUT_DIR + '\\debug.txt'

# Input Excel Cell Position (1 Origin)
IN_CELL_ROW_OFFSET = 4
IN_CELL_COL_OFFSET = 2

# Output Excel Cell Position (1 Origin)
OUT_CELL_ROW_OFFSET = 4
OUT_CELL_COL_OFFSET = 2

# Output Excel Cell Format
ALIGN_LEFT = Alignment(horizontal='left', vertical='top', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='top', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='top', wrap_text=True)
FONT_MEIRYO = Font(name='Meiryo UI', size=10, color='000000')
FONT_MEIRYO_GRAY = Font(name='Meiryo UI', size=10, color='C0C0C0')
FONT_MEIRYO_BOLD = Font(name='Meiryo UI', size=10, color='000000', bold=True)
FILL_BRIGHT_GRAY = PatternFill(patternType='solid', fgColor='EBECF0')
NUMBER_FORMAT_CUBE_INT = '#,##0_ '
NUMBER_FORMAT_CUBE_CENT = '#,##0.00_ '
NUMBER_FORMAT_NO_CUBE_INT = '0_ '
NUMBER_FORMAT_NO_CUBE_CENT = '0.00_ '
BORDER_ALL = Border(
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'),
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'))

# Output Excel Cell Formats
CELL_FORMATS = [
    {'font': FONT_MEIRYO, 'alignment': None,       'number_format': NUMBER_FORMAT_NO_CUBE_INT},
    {'font': FONT_MEIRYO, 'alignment': ALIGN_LEFT, 'number_format': None},
    {'font': FONT_MEIRYO, 'alignment': None,       'number_format': NUMBER_FORMAT_CUBE_CENT},
    {'font': FONT_MEIRYO, 'alignment': None,       'number_format': NUMBER_FORMAT_CUBE_INT},
]


# Read Excel
class ReadExcel:

    def __init__(self, in_excel: str, in_sheet: str) -> None:
        self._wb = openpyxl.load_workbook(in_excel)
        self._sheet = self._wb[in_sheet]
        self._row_offset = IN_CELL_ROW_OFFSET
        self._row = 0
        return

    def range(self) -> (int, int, int, int):
        return IN_CELL_ROW_OFFSET, self._sheet.max_row + 1, IN_CELL_COL_OFFSET, self._sheet.max_column + 1

    def next_row(self) -> None:
        self._row += 1
        return

    def cell(self, i_col: int):
        return self._sheet.cell(row=self._row_offset + self._row, column=i_col)

    def close(self) -> None:
        self._wb.close()
        return


# Write Excel
class WriteExcel:

    def __init__(self, in_excel: str, out_excel: str, out_sheet: str) -> None:
        shutil.copy(in_excel, out_excel)
        self._wb = openpyxl.load_workbook(out_excel)
        self._sheet = self._wb[out_sheet]
        self._row_offset = OUT_CELL_ROW_OFFSET
        self._row = 0
        self._out_excel = out_excel
        return

    def next_row(self) -> None:
        self._row += 1
        return

    def cell(self, i_col: int):
        return self._sheet.cell(row=self._row_offset + self._row, column=i_col)

    def close(self) -> None:
        self._wb.save(self._out_excel)
        self._wb.close()
        return


# Scan Excel File
def scan_excel_file(write_excel: WriteExcel, full_path_file: str,  fp) -> int:

    read_excel = ReadExcel(full_path_file, IN_SHEET)
    row_min, row_max, col_min, col_max = read_excel.range()

    num_lines = 0

    # Read Rows
    for row_current in range(row_min, row_max):

        num_lines += 1
        sep = ''
        csv = ''
        idx = 0

        # Read And Write Cells
        for col_current in range(col_min, col_max):

            # Write Cell
            from_cell = read_excel.cell(col_current)
            to_cell = write_excel.cell(col_current)
            to_cell.value = from_cell.value
            to_cell.border = BORDER_ALL
            to_cell.font = CELL_FORMATS[idx]['font']
            if CELL_FORMATS[idx]['alignment'] is not None:
                to_cell.alignment = CELL_FORMATS[idx]['alignment']
            if CELL_FORMATS[idx]['number_format'] is not None:
                to_cell.number_format = CELL_FORMATS[idx]['number_format']

            csv += sep + ('%s' % to_cell.value)
            sep = ', '
            idx += 1

        read_excel.next_row()
        write_excel.next_row()
        if fp is not None:
            fp.write('%5d: %s\n' % (num_lines, csv))

    read_excel.close()

    return num_lines


# Seek Directories
def seek_directories(excel: WriteExcel, level: int, dir_root: str, dir_relative: str, fp) -> None:

    dirs = []
    files = []

    for path in os.listdir(dir_root):
        if os.path.isfile(os.path.join(dir_root, path)):
            files.append(path)
        else:
            dirs.append(path)

    files.sort(key=str.lower)
    for file in files:
        full_path_file = os.path.join(dir_root, file)
        if fp is not None:
            fp.write('%s\n' % full_path_file)
        base, ext = os.path.splitext(file)
        if ext == EXTEND and not base.startswith('~'):
            lines = scan_excel_file(excel, full_path_file, fp)
            print('%s %s %d' % (dir_relative, file, lines))

    dirs.sort(key=str.lower)
    for dir_nest in dirs:
        seek_directories(excel, level + 1,
                         os.path.join(dir_root, dir_nest), os.path.join(dir_relative, dir_nest), fp)

    return


# Get Current Time
def get_current_time() -> str:

    now = datetime.datetime.now()
    dt = now.strftime("%Y-%m-%d %H:%M:%S")
    return dt


# Main
def main() -> None:

    try:
        options, arguments = getopt.getopt(sys.argv[1:], shortopts="h", longopts=["help"])
    except getopt.error as message:
        print(message)
        print(__doc__)
        sys.exit(1)

    for option, argument in options:
        if option in ("-h", "--help"):
            print(__doc__)
            sys.exit(0)

    print('Excel Summarizer - start [%s]' % get_current_time())

    # fp = None
    fp = open(OUT_DEBUG, 'w', encoding='utf-8')
    write_excel = WriteExcel(IN_EXCEL, OUT_EXCEL, OUT_SHEET)

    seek_directories(write_excel, 0, IN_SRC_ROOT + IN_SRC_RELATIVE, IN_SRC_RELATIVE, fp)

    write_excel.close()
    if fp is not None:
        fp.close()

    print('Excel Summarizer - end [%s]' % get_current_time())

    sys.exit(0)


# Goto Main
if __name__ == '__main__':
    main()
